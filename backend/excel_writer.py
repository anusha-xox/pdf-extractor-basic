"""
excel_writer.py
Writes extracted debit-memo data from multiple PDFs into a single .xlsx file.

Sheet layout
------------
Summary  – one row per PDF with all header-level fields.
Line Items – all line items from all PDFs, with a `source_file` column.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

SUMMARY_COLUMNS = [
    "source_file",
    "debit_memo_number",
    "debit_memo_date",
    "vendor_name",
    "vendor_id",
    "po_number",
    "invoice_number",
    "invoice_date",
    "currency",
    "subtotal",
    "tax",
    "total_amount",
    "reason",
]

LINE_ITEM_COLUMNS = [
    "source_file",
    "debit_memo_number",
    "line_no",
    "description",
    "quantity",
    "unit_price",
    "line_total",
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL     = PatternFill("solid", fgColor="EFF3FB")   # light blue-grey


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill  = _HEADER_FILL
        cell.font  = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)


def _freeze_and_filter(ws, num_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_excel(results: list[tuple[str, dict]], output_path: str) -> None:
    """
    Parameters
    ----------
    results      : list of (pdf_filename, extracted_fields_dict)
    output_path  : destination .xlsx path
    """
    wb = Workbook()
    ws_summary   = wb.active
    ws_summary.title = "Summary"
    ws_line_items = wb.create_sheet("Line Items")

    # Write headers
    ws_summary.append([c.replace("_", " ").title() for c in SUMMARY_COLUMNS])
    ws_line_items.append([c.replace("_", " ").title() for c in LINE_ITEM_COLUMNS])

    for row_idx, (pdf_file, fields) in enumerate(results, start=2):
        source = os.path.basename(pdf_file)
        memo_no = fields.get("debit_memo_number")

        # Summary row
        summary_row = [fields.get(col) if col != "source_file" else source
                       for col in SUMMARY_COLUMNS]
        ws_summary.append(summary_row)

        # Alternate row shading
        if row_idx % 2 == 0:
            for col in range(1, len(SUMMARY_COLUMNS) + 1):
                ws_summary.cell(row=row_idx, column=col).fill = _ALT_FILL

        # Line-item rows
        for item in fields.get("line_items", []):
            li_row = []
            for col in LINE_ITEM_COLUMNS:
                if col == "source_file":
                    li_row.append(source)
                elif col == "debit_memo_number":
                    li_row.append(memo_no)
                else:
                    li_row.append(item.get(col))
            ws_line_items.append(li_row)

    # Apply styles
    for ws, cols in [(ws_summary, SUMMARY_COLUMNS), (ws_line_items, LINE_ITEM_COLUMNS)]:
        _style_header_row(ws, len(cols))
        _auto_width(ws)
        _freeze_and_filter(ws, len(cols))

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")
